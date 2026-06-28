# -*- coding: utf-8 -*-
"""
식단표 내보내기/인쇄 유틸
==========================
- 텍스트 저장(.txt)
- 엑셀 저장(.xlsx) — 가게에서 쓰던 서식 그대로:
    · 배달형(preset='달력형') : 세로 A4, 한 시트에 [중식][석식] 월간 달력.
        제목 Arial 60 / 요일헤더 Arial 29굵게(F3F3F3) / 날짜 GoogleSansText 29굵게
        / 메뉴 GoogleSansText 26(6개를 빈 줄로 연결) 중식=D9D2E9·석식=FCE5CD, 빈칸=D0E0E3.
    · 주방형(preset='주방형') : 가로 A4, 주 단위 4행 블록(요일/날짜/중식/석식).
        구분=Arial 24굵게, 중식=D9D2E9·석식=FCE5CD, 날짜헤더=CFE2F3, 메뉴=F3F3F3.
- 인쇄(Windows: 메모장 인쇄 동사 / 그 외 안내)
"""

from __future__ import annotations

import datetime
import os
import sys
import tempfile

from menu_planner import parse_menu_text
import layouts
from layouts import (
    PRESET_KITCHEN, SECTIONS, ITEMS_PER_MEAL, WEEKDAYS_KR,
    FOOTER_NOTE, MonthData, build_month_data, calendar_weeks,
)

# ---- 가게 서식 상수(7월식단표.xlsx 기준) ----
FONT_HEAD = "Arial"
FONT_BODY = "Google Sans Text"
TXT = "FF1F1F1F"          # 본문 글자색(ARGB)
GRAY = "FFF3F3F3"         # 요일/날짜 헤더, 주방 메뉴
LAVENDER = "FFD9D2E9"     # 중식
PEACH = "FFFCE5CD"        # 석식
TEAL = "FFD0E0E3"         # 배달 빈칸
BLUE = "FFCFE2F3"         # 주방 날짜헤더
FOOTER_TEXT = "\n" + FOOTER_NOTE
WD = WEEKDAYS_KR          # 일~토


def save_txt(text: str, path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        f.write(text)


# ----------------------------------------------------------------------------
# 엑셀 저장 (.xlsx)
# ----------------------------------------------------------------------------

def save_xlsx(text: str, path: str, title: str = "식단표",
              preset: str = layouts.PRESET_CALENDAR,
              year: int | None = None, month: int | None = None) -> None:
    from openpyxl import Workbook

    year, month = _infer_year_month(text, year, month)
    md = build_month_data(text, year, month)
    wb = Workbook()
    wb.remove(wb.active)
    if preset == PRESET_KITCHEN:
        _xlsx_kitchen(wb, md)
    else:
        _xlsx_delivery(wb, md)
    wb.save(path)


def _infer_year_month(text, year, month):
    if year and month:
        return year, month
    mo = None
    for meal in parse_menu_text(text):
        p = layouts.parse_date_label(meal.date_label)
        if p:
            mo = p[0]
            break
    today = datetime.date.today()
    return year or today.year, month or mo or today.month


# ---- 셀 스타일 헬퍼 ----
def _border():
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _set(ws, r, c, value, *, font, size, bold=False, color=TXT, fill=None,
         halign="center", valign="center", wrap=False, border=None):
    from openpyxl.styles import Font, Alignment, PatternFill
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = Font(name=font, size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = border if border is not None else _border()
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def _merge_box(ws, r1, c1, r2, c2, value, *, font, size, bold=False, color=TXT,
               fill=None, halign="center", valign="center"):
    """범위 병합 + 바깥 테두리(모든 셀에 테두리)."""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    bd = _border()
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.border = bd
            if fill:
                from openpyxl.styles import PatternFill
                cell.fill = PatternFill("solid", fgColor=fill)
    _set(ws, r1, c1, value, font=font, size=size, bold=bold, color=color,
         fill=fill, halign=halign, valign=valign)


def _page(ws, landscape):
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_margins = PageMargins(left=0 if landscape else 0.7,
                                  right=0 if landscape else 0.7,
                                  top=0.75, bottom=0.75)


# ---- 배달형(달력형): 세로 A4, 한 시트에 중식/석식 ----
def _xlsx_delivery(wb, md: MonthData) -> None:
    ws = wb.create_sheet("배달")
    ws.column_dimensions["A"].width = 31.4
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 13.0
    _page(ws, landscape=False)

    weeks = calendar_weeks(md.year, md.month)
    r = 1
    for section in SECTIONS:
        r = _delivery_section(ws, md, section, weeks, r)


def _delivery_section(ws, md, section, weeks, r):
    menu_fill = LAVENDER if section == "중식" else PEACH

    # 제목 (3행 병합, Arial 60)
    _merge_box(ws, r, 1, r + 2, 7, f"{md.month}월 식단표 ({section})",
               font=FONT_HEAD, size=60, color=None)
    for k in range(3):
        ws.row_dimensions[r + k].height = 37.5
    r += 3

    # 요일 헤더 (Arial 29 굵게, F3F3F3)
    for ci, wd in enumerate(WD, start=1):
        _set(ws, r, ci, wd, font=FONT_HEAD, size=29, bold=True, color=None,
             fill=GRAY, valign="center")
    ws.row_dimensions[r].height = 37.5
    r += 1

    # 주별: 날짜행 + 메뉴행
    for week in weeks:
        for ci, day in enumerate(week, start=1):
            _set(ws, r, ci, (day or None), font=FONT_BODY, size=29, bold=True,
                 fill=(GRAY if day else TEAL), valign="center")
        ws.row_dimensions[r].height = 37.5
        for ci, day in enumerate(week, start=1):
            has = day and md.has_meal(day, section)
            val = "\n\n".join(md.get(day, section)) if has else None
            _set(ws, r + 1, ci, val, font=FONT_BODY, size=26,
                 fill=(menu_fill if day else TEAL), valign="top")
        ws.row_dimensions[r + 1].height = 337.5
        r += 2

    # 안내문 (3행 병합, Arial 29 굵게, 정렬 일반=왼쪽)
    _merge_box(ws, r, 1, r + 2, 7, FOOTER_TEXT, font=FONT_HEAD, size=29,
               bold=True, color=None, halign=None, valign="top")
    for k in range(3):
        ws.row_dimensions[r + k].height = 37.5
    return r + 3


# ---- 주방형: 가로 A4, 주 단위 4행 블록 ----
def _xlsx_kitchen(wb, md: MonthData) -> None:
    ws = wb.create_sheet("주방")
    ws.column_dimensions["A"].width = 13.0
    ws.column_dimensions["B"].width = 25.1
    for c in "CDEFGH":
        ws.column_dimensions[c].width = 13.0
    _page(ws, landscape=True)

    weeks = calendar_weeks(md.year, md.month)
    r = 1
    for week in weeks:
        if not any(d and md.meals.get(d) for d in week):
            continue
        # 헤더행 + 날짜행: 구분 칸은 두 행 병합
        _merge_box(ws, r, 1, r + 1, 1, "구분", font=FONT_HEAD, size=24, bold=True,
                   color=None, fill=GRAY)
        for ci, wd in enumerate(WD):
            _set(ws, r, ci + 2, wd, font=FONT_HEAD, size=24, bold=True, color=None,
                 fill=BLUE, valign="center")
            _set(ws, r + 1, ci + 2, (week[ci] or None), font=FONT_BODY, size=24,
                 bold=True, fill=BLUE, valign="top")
        ws.row_dimensions[r].height = 33.8
        ws.row_dimensions[r + 1].height = 33.8

        # 중식/석식 행
        for k, section in enumerate(SECTIONS):
            rr = r + 2 + k
            _set(ws, rr, 1, section, font=FONT_BODY, size=24, bold=True,
                 fill=(LAVENDER if section == "중식" else PEACH), valign="center")
            for ci, day in enumerate(week):
                has = day and md.has_meal(day, section)
                val = "\n\n".join(md.get(day, section)) if has else None
                _set(ws, rr, ci + 2, val, font=FONT_BODY, size=24, fill=GRAY,
                     valign="top")
            ws.row_dimensions[rr].height = 311.2
        r += 4


# ----------------------------------------------------------------------------
# 인쇄
# ----------------------------------------------------------------------------

def print_text(text: str, title: str = "식단표") -> str:
    body = f"{title}\n{'=' * 40}\n\n{text}\n"
    fd, path = tempfile.mkstemp(suffix=".txt", prefix="menu_")
    with os.fdopen(fd, "w", encoding="utf-8") as f:
        f.write(body)
    if sys.platform.startswith("win"):
        os.startfile(path, "print")  # type: ignore[attr-defined]
    return path
